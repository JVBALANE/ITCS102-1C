import os
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

def clear_boxes():
    custname_entry.delete(0, tk.END)
    contact_entry.delete(0, tk.END)
    vehicletype_entry.set("")
    plate_entry.delete(0, tk.END)
    startdate_entry.delete(0, tk.END)
    enddate_entry.delete(0, tk.END)
    dailyrate_entry.delete(0, tk.END)
    status_entry.set("Reserved")
    days_label.config(text="0")
    total_label.config(text="0")


def delete_record():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to delete.")
        return

    confirm = messagebox.askyesno("Confirm", "Delete this record?")
    if not confirm:
        return

    values = table.item(selected, "values")
    delete_id = str(values[0])

    wb = load_workbook("Balane_Database.xlsx")
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == delete_id:
            ws.delete_rows(row)
            break
    wb.save("Balane_Database.xlsx")
    wb.close()

    table.delete(selected)
    clear_boxes()
    messagebox.showinfo("Success", "Record deleted.")


def set_rate(event):
    vehicle = vehicletype_entry.get()
    if vehicle == "Motorcycle":
        dailyrate_entry.delete(0, tk.END)
        dailyrate_entry.insert(0, "40")
    elif vehicle == "Sedan":
        dailyrate_entry.delete(0, tk.END)
        dailyrate_entry.insert(0, "100")
    elif vehicle == "SUV":
        dailyrate_entry.delete(0, tk.END)
        dailyrate_entry.insert(0, "120")
    elif vehicle == "Truck":
        dailyrate_entry.delete(0, tk.END)
        dailyrate_entry.insert(0, "300")

def calculate_cost():
    start_txt = startdate_entry.get()
    end_txt = enddate_entry.get()
    rate_txt = dailyrate_entry.get()

    if not start_txt or not end_txt or not rate_txt.isdigit():
        messagebox.showerror("Error", "Fill dates first.")
        return

    try:
        start = datetime.strptime(start_txt, "%Y-%m-%d")
        end = datetime.strptime(end_txt, "%Y-%m-%d")
        if end < start:
            messagebox.showerror("Error", "End date cannot be earlier.")
            return

        days = (end - start).days + 1
        total = days * int(rate_txt)

        days_label.config(text=str(days))
        total_label.config(text=str(total))

    except:
        messagebox.showerror("Error", "Use date format: YYYY-MM-DD")


def submit():
    name = custname_entry.get()
    contact = contact_entry.get()
    vehicle = vehicletype_entry.get()
    plate = plate_entry.get()
    start_dt = startdate_entry.get()
    end_dt = enddate_entry.get()
    rate = dailyrate_entry.get()
    status = status_entry.get()
    days = days_label.cget("text")
    total = total_label.cget("text")

    if not name or not contact or not vehicle or not plate or not start_dt or not end_dt or days == "0":
        messagebox.showerror("Error", "All fields required. Calculate first.")
        return

    if not contact.isdigit():
        messagebox.showerror("Error", "Contact must be numbers.")
        return

    if os.path.exists("Balane_Database.xlsx"):
        wb = load_workbook("Balane_Database.xlsx")
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "CustomerName", "Contact", "VehicleType", "Plate",
                   "StartDate", "EndDate", "Days", "Rate", "Total", "Status"])

    new_id = f"RES{ws.max_row}"
    ws.append([new_id, name, contact, vehicle, plate, start_dt, end_dt, days, rate, total, status])
    wb.save("Balane_Database.xlsx")
    wb.close()

    messagebox.showinfo("Success", "Saved successfully!")
    clear_boxes()
    load_data()


def change_record():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Select record first.")
        return

    name = custname_entry.get()
    contact = contact_entry.get()
    vehicle = vehicletype_entry.get()
    plate = plate_entry.get()
    start_dt = startdate_entry.get()
    end_dt = enddate_entry.get()
    rate = dailyrate_entry.get()
    status = status_entry.get()
    days = days_label.cget("text")
    total = total_label.cget("text")

    if not name or not contact or not vehicle or not plate or not start_dt or not end_dt or days == "0":
        messagebox.showerror("Error", "All fields required. Calculate first.")
        return

    values = table.item(selected, "values")
    record_id = str(values[0])

    wb = load_workbook("Balane_Database.xlsx")
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == record_id:
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=contact)
            ws.cell(row=row, column=4, value=vehicle)
            ws.cell(row=row, column=5, value=plate)
            ws.cell(row=row, column=6, value=start_dt)
            ws.cell(row=row, column=7, value=end_dt)
            ws.cell(row=row, column=8, value=days)
            ws.cell(row=row, column=9, value=rate)
            ws.cell(row=row, column=10, value=total)
            ws.cell(row=row, column=11, value=status)
            break
    wb.save("Balane_Database.xlsx")
    wb.close()

    messagebox.showinfo("Success", "Record updated.")
    load_data()
    clear_boxes()


def select_record(event):
    selected = table.focus()
    if not selected:
        return
    values = table.item(selected, "values")
    clear_boxes()

    custname_entry.insert(0, values[1])
    contact_entry.insert(0, values[2])
    vehicletype_entry.set(values[3])
    plate_entry.insert(0, values[4])
    startdate_entry.insert(0, values[5])
    enddate_entry.insert(0, values[6])
    dailyrate_entry.insert(0, values[8])
    status_entry.set(values[10])
    days_label.config(text=values[7])
    total_label.config(text=values[9])

def load_data():
    table.delete(*table.get_children())
    if not os.path.exists("Balane_Database.xlsx"):
        return
    wb = load_workbook("Balane_Database.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            table.insert("", "end", values=row)
    wb.close()

window = tk.Tk()
window.title("Vehicle Reservation System")
window.geometry("900x580")
window.resizable(False, False)

title = tk.Label(window, text="Vehicle Reservation System", font=("Arial", 16))
title.pack(pady=10)

frame = tk.Frame(window)
frame.pack(pady=5)

tk.Label(frame, text="Customer Name").grid(row=0, column=0, padx=10, pady=8)
custname_entry = tk.Entry(frame, width=25)
custname_entry.grid(row=0, column=1)

tk.Label(frame, text="Contact Number").grid(row=0, column=2, padx=10, pady=8)
contact_entry = tk.Entry(frame, width=25)
contact_entry.grid(row=0, column=3)

tk.Label(frame, text="Vehicle Type").grid(row=1, column=0, padx=10, pady=8)
vehicletype_entry = ttk.Combobox(frame, values=["Motorcycle", "Sedan", "SUV", "Truck"], state="readonly", width=22)
vehicletype_entry.grid(row=1, column=1)
vehicletype_entry.bind("<<ComboboxSelected>>", set_rate)

tk.Label(frame, text="Plate Number").grid(row=1, column=2, padx=10, pady=8)
plate_entry = tk.Entry(frame, width=25)
plate_entry.grid(row=1, column=3)

tk.Label(frame, text="Start Date (YYYY-MM-DD)").grid(row=2, column=0, padx=10, pady=8)
startdate_entry = tk.Entry(frame, width=25)
startdate_entry.grid(row=2, column=1)

tk.Label(frame, text="End Date (YYYY-MM-DD)").grid(row=2, column=2, padx=10, pady=8)
enddate_entry = tk.Entry(frame, width=25)
enddate_entry.grid(row=2, column=3)

tk.Label(frame, text="Daily Rate").grid(row=3, column=0, padx=10, pady=8)
dailyrate_entry = tk.Entry(frame, width=25)
dailyrate_entry.grid(row=3, column=1)

tk.Label(frame, text="Status").grid(row=3, column=2, padx=10, pady=8)
status_entry = ttk.Combobox(frame, values=["Reserved", "Ongoing", "Completed", "Cancelled"], state="readonly", width=22)
status_entry.set("Reserved")
status_entry.grid(row=3, column=3)

tk.Label(frame, text="Days Rented:").grid(row=4, column=0, padx=10, pady=8)
days_label = tk.Label(frame, text="0")
days_label.grid(row=4, column=1)

tk.Label(frame, text="Total Cost:").grid(row=4, column=2, padx=10, pady=8)
total_label = tk.Label(frame, text="0")
total_label.grid(row=4, column=3)

tk.Button(frame, text="Calculate", command=calculate_cost, bg="skyblue", fg="white").grid(row=4, column=4, padx=5)

tk.Button(frame, text="Submit", command=submit, bg="green", fg="white", width=10).grid(row=5, column=0, padx=5, pady=12)
tk.Button(frame, text="Update", command=change_record, bg="orange", fg="white", width=10).grid(row=5, column=1, padx=5)
tk.Button(frame, text="Delete", command=delete_record, bg="red", fg="white", width=10).grid(row=5, column=2, padx=5)
tk.Button(frame, text="Clear", command=clear_boxes, bg="gray", fg="white", width=10).grid(row=5, column=3, padx=5)

table_frame = tk.Frame(window)
table_frame.pack(fill="both", expand=True, padx=10, pady=5)

columns = ("ID", "Name", "Contact", "Vehicle", "Plate", "Start", "End", "Days", "Rate", "Total", "Status")
table = ttk.Treeview(table_frame, columns=columns, show="headings")

for col in columns:
    table.heading(col, text=col)
    table.column(col, width=75, anchor="center")

table.pack(fill="both", expand=True)
table.bind("<ButtonRelease-1>", select_record)

load_data()
window.mainloop()
